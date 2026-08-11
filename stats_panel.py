"""
作业上传数据统计面板模块
功能: 提供数据报告、柱状图、折线图、上传记录表、失败记录表及Excel导出
技术: customtkinter + tkinter.Canvas + openpyxl (卡片化布局 + 扁平化图表)
"""
import os
import tkinter as tk
import tkinter.font as tkfont
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta

import customtkinter as ctk

import ui_theme as theme
import win32_helpers


class StatsPanel:
    """
    数据统计面板
    包含图表可视化、数据表格和Excel导出功能
    """

    def __init__(self, parent, db, root):
        self.parent = parent
        self.db = db
        self.root = root

        # 图表状态
        self._bar_mode = "subject"       # "subject" | "school_grade"
        self._line_mode = "daily"        # "daily" | "weekly" | "monthly"

        # 图表画布与重绘状态
        self._bar_canvas = None
        self._line_canvas = None
        self._bar_after_id = None
        self._line_after_id = None
        # 页面隐藏时 winfo_width() 返回1,用上次有效宽度兜底(默认800)
        self._chart_widths = {"bar": 800, "line": 800}
        self._measure_fonts = {}

        # 构建界面
        self._create_widgets()

    # ==================== 界面构建 ====================

    def _create_widgets(self):
        """创建统计面板所有界面组件(纵向滚动卡片流)"""
        self._scroll = ctk.CTkScrollableFrame(
            self.parent, fg_color="transparent",
            scrollbar_button_color=theme.TEXT_FAINT,
            scrollbar_button_hover_color=theme.TEXT_MUTED)
        self._scroll.pack(fill="both", expand=True)

        # === 1. 数据报告区 ===
        self._create_action_bar()

        # === 2. 上传记录表 ===
        self._create_upload_table_section()

        # === 3. 柱状图: 作业上传数量 ===
        self._create_bar_chart_section()

        # === 4. 折线图: 作业上传趋势 ===
        self._create_line_chart_section()

        # === 5. 失败记录表 ===
        self._create_failed_table_section()

        # 初始加载表格数据
        self._refresh_upload_table()
        self._refresh_failed_table()

        # 修复滚动残影(在表格创建完成后绑定)
        self._bind_scroll_ghost_fix()

    def _new_card(self, title: str, subtitle: str = None):
        """创建一张带标题的卡片并返回卡片容器"""
        card = theme.card(self._scroll)
        card.pack(fill="x", pady=(0, 12))

        head = ctk.CTkFrame(card, fg_color="transparent")
        head.pack(fill="x", padx=20, pady=(16, 8))
        theme.card_title(head, title).pack(side="left")
        if subtitle:
            ctk.CTkLabel(head, text=subtitle, font=theme.font(11),
                         text_color=theme.TEXT_FAINT).pack(side="right")
        return card

    def _create_action_bar(self):
        """数据报告区: 状态提示 + 分析报告按钮"""
        card = self._new_card("数据报告",
                              "上传成功自动同步到分析表，数据持久保留")

        btn_row = ctk.CTkFrame(card, fg_color="transparent")
        btn_row.pack(fill="x", padx=20, pady=(0, 18))

        for label, period in [("今日报告", "daily"), ("近7天报告", "weekly"),
                              ("近30天报告", "monthly")]:
            theme.ghost_button(btn_row, label,
                               lambda p=period: self._generate_report(p),
                               width=110).pack(side="left", padx=(0, 10))

        theme.primary_button(btn_row, "打开报告目录", self._open_reports_dir,
                             width=120).pack(side="right")

    def _generate_report(self, period: str):
        """
        生成失败分析报告

        Args:
            period: 'daily' 今日 / 'weekly' 近7天 / 'monthly' 近30天
        """
        # 延迟导入避免循环依赖
        from failure_analysis_agent import FailureAnalysisAgent

        today = datetime.now()
        if period == 'daily':
            start = today.strftime('%Y-%m-%d')
            end = today.strftime('%Y-%m-%d')
            report_type = 'daily'
        elif period == 'weekly':
            start = (today - timedelta(days=7)).strftime('%Y-%m-%d')
            end = today.strftime('%Y-%m-%d')
            report_type = 'weekly'
        elif period == 'monthly':
            start = (today - timedelta(days=30)).strftime('%Y-%m-%d')
            end = today.strftime('%Y-%m-%d')
            report_type = 'custom'
        else:
            return

        try:
            agent = FailureAnalysisAgent()
            path = agent.generate_report(start, end, report_type)
            if path:
                messagebox.showinfo("报告生成成功",
                                    f"分析报告已生成：\n{path}\n\n是否打开报告？")
                os.startfile(path)
            else:
                messagebox.showwarning("报告生成失败",
                                       "报告中无数据或生成过程出错，请检查日志")
        except Exception as e:
            messagebox.showerror("错误", f"生成报告时发生异常：\n{e}")

    def _open_reports_dir(self):
        """打开报告目录"""
        from failure_analysis_agent import FailureAnalysisAgent
        try:
            agent = FailureAnalysisAgent()
            path = agent.open_reports_dir()
        except Exception as e:
            messagebox.showerror("错误", f"打开报告目录失败：\n{e}")

    def _create_bar_chart_section(self):
        """柱状图区: 作业上传数量图"""
        card = self._new_card("上传数量统计")

        self._bar_segment = ctk.CTkSegmentedButton(
            card, values=["按科目", "按学校+年级"], font=theme.font(11),
            height=30, corner_radius=8,
            fg_color=theme.CARD_INNER,
            selected_color=theme.PRIMARY, selected_hover_color=theme.PRIMARY_HOVER,
            unselected_color=theme.CARD_INNER, unselected_hover_color=theme.PRIMARY_SOFT,
            text_color=theme.TEXT_MUTED,
            command=self._on_bar_segment_change)
        self._bar_segment.set("按科目")
        self._bar_segment.pack(anchor="w", padx=20, pady=(0, 8))

        self._bar_canvas = tk.Canvas(card, height=350, bg=theme.CARD,
                                     bd=0, highlightthickness=0)
        self._bar_canvas.pack(fill="x", padx=20, pady=(0, 18))
        self._bar_canvas.bind("<Configure>", self._on_bar_resize)

        # 初始尝试加载数据
        self._refresh_bar_chart()

    def _create_line_chart_section(self):
        """折线图区: 作业上传趋势图"""
        card = self._new_card("上传趋势")

        self._line_segment = ctk.CTkSegmentedButton(
            card, values=["按日", "按周", "按月"], font=theme.font(11),
            height=30, corner_radius=8,
            fg_color=theme.CARD_INNER,
            selected_color=theme.PRIMARY, selected_hover_color=theme.PRIMARY_HOVER,
            unselected_color=theme.CARD_INNER, unselected_hover_color=theme.PRIMARY_SOFT,
            text_color=theme.TEXT_MUTED,
            command=self._on_line_segment_change)
        self._line_segment.set("按日")
        self._line_segment.pack(anchor="w", padx=20, pady=(0, 8))

        self._line_canvas = tk.Canvas(card, height=350, bg=theme.CARD,
                                      bd=0, highlightthickness=0)
        self._line_canvas.pack(fill="x", padx=20, pady=(0, 18))
        self._line_canvas.bind("<Configure>", self._on_line_resize)

        self._refresh_line_chart()

    def _create_upload_table_section(self):
        """上传记录表(支持学校/年级/科目筛选，可勾选是否完成布置)"""
        card = self._new_card("上传记录", "点击行末方框可勾选是否完成布置")

        # 筛选行: 学校/年级/科目下拉框 + 导出按钮
        btn_row = ctk.CTkFrame(card, fg_color="transparent")
        btn_row.pack(fill="x", padx=20, pady=(0, 8))

        self._filter_combos = {}
        for field, label in (("school", "学校"), ("grade", "年级"), ("subject", "科目")):
            ctk.CTkLabel(btn_row, text=label, font=theme.font(12),
                         text_color=theme.TEXT_MUTED).pack(side="left", padx=(0, 6))
            combo = ctk.CTkComboBox(
                btn_row, width=120, height=30, corner_radius=8, state="readonly",
                font=theme.font(12), dropdown_font=theme.font(12),
                values=["全部"],
                fg_color=theme.CARD_INNER, border_color=theme.BORDER,
                button_color=theme.CARD_INNER, button_hover_color=theme.PRIMARY_SOFT,
                text_color=theme.TEXT, dropdown_fg_color=theme.CARD,
                dropdown_hover_color=theme.PRIMARY_SOFT,
                command=lambda _v: self._refresh_upload_table())
            combo.set("全部")
            combo.pack(side="left", padx=(0, 14))
            self._filter_combos[field] = combo

        # 是否完成布置筛选(固定三项,不参与动态选项刷新)
        ctk.CTkLabel(btn_row, text="布置", font=theme.font(12),
                     text_color=theme.TEXT_MUTED).pack(side="left", padx=(0, 6))
        self._assigned_combo = ctk.CTkComboBox(
            btn_row, width=100, height=30, corner_radius=8, state="readonly",
            font=theme.font(12), dropdown_font=theme.font(12),
            values=["全部", "已完成", "未完成"],
            fg_color=theme.CARD_INNER, border_color=theme.BORDER,
            button_color=theme.CARD_INNER, button_hover_color=theme.PRIMARY_SOFT,
            text_color=theme.TEXT, dropdown_fg_color=theme.CARD,
            dropdown_hover_color=theme.PRIMARY_SOFT,
            command=lambda _v: self._refresh_upload_table())
        self._assigned_combo.set("全部")
        self._assigned_combo.pack(side="left", padx=(0, 14))

        theme.ghost_button(btn_row, "导出到 Excel",
                           self._export_upload_table,
                           width=110).pack(side="right")

        columns = ("file_name", "school", "grade", "subject", "upload_time", "assigned")
        headers = ("作业名称", "学校", "年级", "科目", "上传时间", "是否完成布置")
        upload_frame, self._upload_tree = self._create_treeview(card, columns, headers, height=8)
        upload_frame.pack(fill="both", expand=True, padx=20, pady=(0, 18))

        # 未勾选行红色、已勾选行绿色
        self._upload_tree.tag_configure("unassigned", foreground=theme.DANGER)
        self._upload_tree.tag_configure("assigned", foreground=theme.SUCCESS)
        # 点击"是否完成布置"列切换勾选状态
        self._upload_tree.bind("<Button-1>", self._on_upload_tree_click)

    def _create_failed_table_section(self):
        """失败记录表"""
        card = self._new_card("失败上传记录")

        btn_row = ctk.CTkFrame(card, fg_color="transparent")
        btn_row.pack(fill="x", padx=20, pady=(0, 8))
        theme.ghost_button(btn_row, "导出到 Excel",
                           self._export_failed_table,
                           width=110).pack(side="right")

        columns = ("file_name", "school", "grade", "subject", "upload_time", "error_message")
        headers = ("作业名称", "学校", "年级", "科目", "上传时间", "失败原因")
        failed_frame, self._failed_tree = self._create_treeview(card, columns, headers, height=8)
        failed_frame.pack(fill="both", expand=True, padx=20, pady=(0, 18))

    @staticmethod
    def _create_treeview(parent, columns, headers, height=10):
        """创建带滚动条的扁平化Treeview，返回 (容器Frame, Treeview) 元组"""
        tree_frame = ctk.CTkFrame(parent, fg_color="transparent")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings",
                            height=height, style="Card.Treeview")

        for col, header in zip(columns, headers):
            tree.heading(col, text=header)
            if col == "error_message":
                width = 200
                anchor = "w"
            elif col == "file_name":
                width = 200
                anchor = "w"
            elif col == "upload_time":
                width = 150
                anchor = "center"
            elif col == "assigned":
                width = 110
                anchor = "center"
            else:
                width = 100
                anchor = "center"
            tree.column(col, width=width, anchor=anchor, minwidth=60)

        sb = theme.attach_tree_scrollbar(tree_frame, tree)
        tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y", padx=(6, 0))

        # 鼠标悬停在表格上时由表格自身滚动，返回"break"阻断事件
        # 传递给外层CTkScrollableFrame，避免整个统计面板跟着滚动
        def _on_mousewheel(event, t=tree):
            first, last = t.yview()
            if first <= 0.0 and last >= 1.0:
                # 表格内容不足一屏，放行事件让外层面板滚动
                return None
            t.yview_scroll(-1 * (event.delta // 120), "units")
            return "break"
        tree.bind("<MouseWheel>", _on_mousewheel)
        return tree_frame, tree

    # ==================== 滚动残影修复 ====================

    def _bind_scroll_ghost_fix(self):
        """
        修复 Windows 下滚动统计面板时表格区域出现残影的问题(委托共享的
        win32_helpers.ScrollGhostFix, 与设置页同一份实现)。

        根因: CTkScrollableFrame 用 Canvas 内嵌真实 HWND Frame，滚动时
        Windows 发送 WM_ERASEBKGND 擦除背景再 WM_PAINT 重绘，与内容帧
        移动不同步产生残影；且 Tk 的重绘是异步的，快速拖动时 WM_PAINT
        被鼠标事件抢占，内容绘制滞后于滑块。
        修复: 1) 用 64 位正确的 GetWindowLongPtrW/SetWindowLongPtrW 子类化
        画布与内容帧窗口过程抑制 WM_ERASEBKGND（旧实现用 32 位
        SetWindowLongW，64 位下静默失败从未生效）；
        2) 滚动事件(拖动中/滚轮)后经 after_idle 合并调度一次 update()
        同步完成重绘(直接调用 update() 会嵌套排空事件队列导致递归重入)。
        失败静默回退，不崩溃。destroy() 时 handle.uninstall() 精确还原。
        """
        self._ghost_fix = win32_helpers.ScrollGhostFix(self._scroll, self.root)
        # 面板被 tkraise 盖住(切到其他页)时跳过重绘, 避免对隐藏页无谓 update()
        self._ghost_fix.active = False

    # ==================== 图表刷新 ====================

    def _on_bar_segment_change(self, value: str):
        """切换柱状图视图: 科目 ↔ 学校+年级"""
        self._bar_mode = "subject" if value == "按科目" else "school_grade"
        self._refresh_bar_chart()

    def _on_line_segment_change(self, value: str):
        """切换折线图聚合方式: daily / weekly / monthly"""
        mapping = {"按日": "daily", "按周": "weekly", "按月": "monthly"}
        self._line_mode = mapping.get(value, "daily")
        self._refresh_line_chart()

    # ==================== Canvas 图表绘制 ====================

    # 绘图区边距(像素): 左(y轴刻度)/右/上(标题)/下(基础,45°x标签时自适应加高)
    _CHART_H = 350
    _M_LEFT = 44
    _M_RIGHT = 12
    _M_TOP = 34
    _M_BOTTOM_BASE = 46

    def _chart_width(self, key: str) -> int:
        """取画布当前宽度;页面隐藏时(winfo_width()==1)用上次有效宽度兜底"""
        cv = self._bar_canvas if key == "bar" else self._line_canvas
        w = cv.winfo_width()
        if w >= 50:
            self._chart_widths[key] = w
        return self._chart_widths[key]

    def _measure(self, text: str, size: int) -> int:
        """测量文本像素宽度(字体按点尺寸与create_text一致,缓存避免反复创建)"""
        f = self._measure_fonts.get(size)
        if f is None:
            f = tkfont.Font(family=theme.FONT_FAMILY, size=size)
            self._measure_fonts[size] = f
        return f.measure(text)

    def _plot_rect(self, width: int, labels):
        """计算绘图区 (x0, x1, y0, y1);底部边距按最长标签自适应(45°旋转水平投影≈0.71×字宽)"""
        max_w = max((self._measure(lbl, 8) for lbl in labels), default=0)
        m_bottom = min(96, max(self._M_BOTTOM_BASE, 30 + int(max_w * 0.71)))
        return self._M_LEFT, width - self._M_RIGHT, self._M_TOP, self._CHART_H - m_bottom

    @staticmethod
    def _nice_y_max(raw: int) -> int:
        """y轴最大值取整(1/2/5×10^n 步进,保证刻度为整数),最小5"""
        if raw <= 0:
            return 5
        exp = 10 ** (len(str(int(raw))) - 1)
        for factor in (1, 2, 5, 10):
            nice = factor * exp
            if nice >= raw:
                return max(nice, 5)
        return max(10 * exp, 5)

    def _draw_axes(self, cv, width, title, labels, y_max):
        """绘制标题、网格线、y轴刻度与"上传数量"标签,返回绘图区坐标"""
        x0, x1, y0, y1 = self._plot_rect(width, labels)
        n_ticks = 5
        # 网格线先画(数据绘制在其上);i==0 的线兼作x轴
        for i in range(n_ticks + 1):
            y = y1 - (i / n_ticks) * (y1 - y0)
            cv.create_line(x0, y, x1, y, fill=theme.BORDER, width=1)
            cv.create_text(x0 - 8, y, text=str(int(round(i * y_max / n_ticks))),
                           anchor="e", font=(theme.FONT_FAMILY, 8), fill=theme.TEXT_MUTED)
        cv.create_line(x0, y0, x0, y1, fill=theme.BORDER, width=1)  # 左侧轴
        cv.create_text(12, (y0 + y1) / 2, text="上传数量", angle=90, anchor="center",
                       font=(theme.FONT_FAMILY, 9), fill=theme.TEXT_MUTED)
        cv.create_text(width / 2, 18, text=title, anchor="center",
                       font=(theme.FONT_FAMILY, 10), fill=theme.TEXT)
        return x0, x1, y0, y1

    @staticmethod
    def _draw_x_label(cv, cx, y1, label):
        """x轴标签45°右对齐(Tk 8.6 canvas text支持angle);若出现镜像文字,改为angle=45, anchor="ne" """
        cv.create_text(cx, y1 + 6, text=label, angle=-45, anchor="se",
                       font=(theme.FONT_FAMILY, 8), fill=theme.TEXT_MUTED)

    def _draw_empty(self, cv, width):
        """空数据: 居中"暂无数据"提示"""
        cv.create_text(width / 2, self._CHART_H / 2, text="暂无数据", anchor="center",
                       font=(theme.FONT_FAMILY, 13), fill=theme.TEXT_FAINT)

    def _on_bar_resize(self, event):
        """窗口尺寸变化防抖重绘(100ms)"""
        if event.width < 50:
            return
        if self._bar_after_id is not None:
            self.root.after_cancel(self._bar_after_id)
        self._bar_after_id = self.root.after(100, self._refresh_bar_chart)

    def _on_line_resize(self, event):
        """窗口尺寸变化防抖重绘(100ms)"""
        if event.width < 50:
            return
        if self._line_after_id is not None:
            self.root.after_cancel(self._line_after_id)
        self._line_after_id = self.root.after(100, self._refresh_line_chart)

    def _refresh_bar_chart(self):
        """重绘柱状图(纯tkinter Canvas)"""
        try:
            if self._bar_mode == "subject":
                data = self.db.get_upload_count_by_subject()
                title = "各科目上传数量统计"
            else:
                data = self.db.get_upload_count_by_school_grade()
                title = "各学校年级上传数量统计"

            cv = self._bar_canvas
            width = self._chart_width("bar")
            cv.delete("all")

            if not data:
                self._draw_empty(cv, width)
                return

            if self._bar_mode == "subject":
                labels = [d["subject"] for d in data]
            else:
                labels = [f"{d['school']}{d['grade']}" for d in data]
            counts = [d["count"] for d in data]

            y_max = self._nice_y_max(max(counts))
            x0, x1, y0, y1 = self._draw_axes(cv, width, title, labels, y_max)

            n = len(labels)
            slot = (x1 - x0) / n
            bar_w = slot * 0.55
            for i, (label, count) in enumerate(zip(labels, counts)):
                cx = x0 + slot * i + slot / 2
                top = y1 - (count / y_max) * (y1 - y0)
                cv.create_rectangle(cx - bar_w / 2, top, cx + bar_w / 2, y1,
                                    fill=theme.PRIMARY, outline="")
                cv.create_text(cx, top - 4, text=str(count), anchor="s",
                               font=(theme.FONT_FAMILY, 8), fill=theme.TEXT_MUTED)
                self._draw_x_label(cv, cx, y1, label)
        except Exception as e:
            print(f"统计面板柱状图绘制失败: {e}")

    def _refresh_line_chart(self):
        """重绘折线图(纯tkinter Canvas)"""
        try:
            data = self.db.get_upload_count_by_date(self._line_mode)

            cv = self._line_canvas
            width = self._chart_width("line")
            cv.delete("all")

            if not data:
                self._draw_empty(cv, width)
                return

            labels = [d["date_label"] for d in data]
            counts = [d["count"] for d in data]
            mode_title = {"daily": "每日", "weekly": "每周", "monthly": "每月"}
            title = f"作业上传趋势 ({mode_title[self._line_mode]})"

            y_max = self._nice_y_max(max(counts))
            x0, x1, y0, y1 = self._draw_axes(cv, width, title, labels, y_max)

            n = len(labels)
            xs = ([x0 + i * (x1 - x0) / (n - 1) for i in range(n)]
                  if n > 1 else [(x0 + x1) / 2])
            ys = [y1 - (c / y_max) * (y1 - y0) for c in counts]

            # 面积填充: PRIMARY_SOFT 恰为 PRIMARY 10%透明度叠加白色
            cv.create_polygon([(x0, y1)] + list(zip(xs, ys)) + [(xs[-1], y1)],
                              fill=theme.PRIMARY_SOFT, outline="")
            # 单点时create_line需至少2个点,跳过(仍显示圆点与数值)
            if n > 1:
                cv.create_line([p for pair in zip(xs, ys) for p in pair],
                               fill=theme.PRIMARY, width=1.6)

            # 标签过多时每隔N个显示一个(与matplotlib版一致)
            step = max(1, n // 20)
            for i, (x, y) in enumerate(zip(xs, ys)):
                cv.create_oval(x - 2, y - 2, x + 2, y + 2,
                               fill=theme.PRIMARY, outline="")
                cv.create_text(x, y - 6, text=str(counts[i]), anchor="s",
                               font=(theme.FONT_FAMILY, 7), fill=theme.TEXT_MUTED)
                if i % step == 0:
                    self._draw_x_label(cv, x, y1, labels[i])
        except Exception as e:
            print(f"统计面板折线图绘制失败: {e}")

    # ==================== 表格刷新 ====================

    # 勾选框字符: □ 未勾选 / ☑ 已勾选
    _CHECK_ON = "☑ 已完成"
    _CHECK_OFF = "□ 未完成"

    def _refresh_filter_options(self):
        """刷新筛选下拉框的可选项(保留当前选中值)"""
        options = self.db.get_analysis_filter_options()
        for field, combo in self._filter_combos.items():
            current = combo.get()
            values = ["全部"] + options.get(field, [])
            combo.configure(values=values)
            if current not in values:
                combo.set("全部")

    def _refresh_upload_table(self):
        """刷新上传记录表(从分析表读取，支持筛选，数据持久不受上传记录清理影响)"""
        # 先刷新筛选项(首次或数据变化时)
        if hasattr(self, "_filter_combos"):
            self._refresh_filter_options()
            filters = {}
            for field, combo in self._filter_combos.items():
                val = combo.get()
                filters[field] = None if val == "全部" else val
        else:
            filters = {"school": None, "grade": None, "subject": None}

        # 是否完成布置筛选: 全部=None / 已完成=True / 未完成=False
        assigned_filter = None
        if hasattr(self, "_assigned_combo"):
            assigned_val = self._assigned_combo.get()
            if assigned_val == "已完成":
                assigned_filter = True
            elif assigned_val == "未完成":
                assigned_filter = False

        for item in self._upload_tree.get_children():
            self._upload_tree.delete(item)
        records = self.db.get_all_analysis_records(
            school=filters["school"], grade=filters["grade"], subject=filters["subject"],
            assigned=assigned_filter)
        for r in records:
            assigned = bool(r.get("assigned"))
            tag = "assigned" if assigned else "unassigned"
            self._upload_tree.insert("", "end", iid=str(r["id"]), tags=(tag,), values=(
                r["file_name"], r["school"], r["grade"],
                r["subject"], r["upload_time"],
                self._CHECK_ON if assigned else self._CHECK_OFF
            ))

    def _on_upload_tree_click(self, event):
        """点击"是否完成布置"列切换勾选状态"""
        tree = self._upload_tree
        if tree.identify_region(event.x, event.y) != "cell":
            return
        if tree.identify_column(event.x) != "#6":  # 第6列 assigned
            return
        item = tree.identify_row(event.y)
        if not item:
            return
        try:
            record_id = int(item)
        except (TypeError, ValueError):
            return
        # 当前是否已勾选
        is_assigned = "assigned" in tree.item(item, "tags")
        new_state = not is_assigned
        self.db.set_analysis_assigned(record_id, new_state)
        # 布置筛选生效时，切换后该行可能不再满足筛选条件，直接刷新表格
        if hasattr(self, "_assigned_combo") and self._assigned_combo.get() != "全部":
            self._refresh_upload_table()
            return "break"
        # 更新行显示
        values = list(tree.item(item, "values"))
        values[5] = self._CHECK_ON if new_state else self._CHECK_OFF
        tree.item(item, values=values,
                  tags=("assigned" if new_state else "unassigned",))
        return "break"

    def _refresh_failed_table(self):
        """刷新失败记录表（从分析表读取，数据持久保留不受上传记录清理影响）"""
        for item in self._failed_tree.get_children():
            self._failed_tree.delete(item)
        records = self.db.get_failed_records_for_stats()
        for r in records:
            self._failed_tree.insert("", "end", values=(
                r["file_name"], r["school"], r["grade"],
                r["subject"], r["upload_time"],
                r["error_message"] or ""
            ))

    # ==================== Excel导出 ====================

    def _export_upload_table(self):
        """导出上传记录表到Excel"""
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")],
            title="导出上传记录表"
        )
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "上传记录"

            # 表头
            headers = ["作业名称", "学校", "年级", "科目", "上传时间", "是否完成布置"]
            header_fill = PatternFill(start_color="5B7CFA", end_color="5B7CFA", fill_type="solid")
            header_font_white = Font(bold=True, size=11, color="FFFFFF")

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # 数据行
            for row_idx, item in enumerate(self._upload_tree.get_children(), 2):
                values = self._upload_tree.item(item, "values")
                for col_idx, val in enumerate(values, 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)

            # 调整列宽
            ws.column_dimensions['A'].width = 40  # 作业名称
            ws.column_dimensions['B'].width = 20  # 学校
            ws.column_dimensions['C'].width = 10  # 年级
            ws.column_dimensions['D'].width = 10  # 科目
            ws.column_dimensions['E'].width = 22  # 上传时间
            ws.column_dimensions['F'].width = 14  # 是否完成布置

            wb.save(path)
            count = len(self._upload_tree.get_children())
            messagebox.showinfo("导出成功", f"已导出 {count} 条上传记录到:\n{path}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    def _export_failed_table(self):
        """导出失败记录表到Excel"""
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")],
            title="导出失败记录表"
        )
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "失败记录"

            headers = ["作业名称", "学校", "年级", "科目", "上传时间", "失败原因", "Agent接管成功"]
            header_fill = PatternFill(start_color="D9695F", end_color="D9695F", fill_type="solid")
            header_font = Font(bold=True, size=11, color="FFFFFF")

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row_idx, item in enumerate(self._failed_tree.get_children(), 2):
                values = self._failed_tree.item(item, "values")
                for col_idx, val in enumerate(values, 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)

            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 22
            ws.column_dimensions['F'].width = 40
            ws.column_dimensions['G'].width = 12

            wb.save(path)
            count = len(self._failed_tree.get_children())
            messagebox.showinfo("导出成功", f"已导出 {count} 条失败记录到:\n{path}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    # ==================== 资源清理 ====================

    def destroy(self):
        """取消待定重绘、解绑事件并清空画布(窗口随后整体销毁)"""
        for key in ("_bar_after_id", "_line_after_id"):
            after_id = getattr(self, key, None)
            if after_id:
                try:
                    self.root.after_cancel(after_id)
                except Exception:
                    pass
                setattr(self, key, None)
        for cv in (self._bar_canvas, self._line_canvas):
            if cv is not None:
                try:
                    cv.unbind("<Configure>")
                    cv.delete("all")
                except Exception:
                    pass
        # 还原子类化窗口过程 + 精确解绑滚动条/全局滚轮绑定
        # (unbind_all(seq, funcid) 只接受一个参数会抛 TypeError, 旧代码
        # 的清理从未真正执行; ScrollGhostFix.uninstall 用 tk.call 精确移除)
        ghost_fix = getattr(self, "_ghost_fix", None)
        if ghost_fix is not None:
            try:
                ghost_fix.uninstall()
            except Exception:
                pass
